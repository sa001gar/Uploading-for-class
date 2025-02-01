import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

def create_excel_file():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activation MS DSR"
    
    # Adding headers
    ws.append(["ACTIVATION MS DSR"])
    ws.append(["MS Name:", "", "", "", "Date:"])
    ws.append(["Route Name:"])
    ws.append([])  # Empty row
    
    # Column Headers
    headers = [
        "Sl. No", "Outlet Name", "Location", "Mobile NO", 
        "Sale", "", "", "", "", "", 
        "Tick (if available)", "", "", "", "", "", 
        "Sampling", "Qty."
    ]
    sub_headers = [
        "", "", "", "", 
        "SES", "SER", "CSF'69", "CSF'64/CRF", "TRL", "CFT", 
        "SES", "SER", "CSF'69", "CSF'64/CRF", "TRL", "CFT", 
        ""
    ]
    ws.append(headers)
    ws.append(sub_headers)
    
    # Merging header cells
    ws.merge_cells("E5:J5")  # Sale
    ws.merge_cells("K5:P5")  # Tick (if available)
    ws.merge_cells("Q5:R5")  # Sampling
    
    # Adding empty rows for data entry
    for i in range(1, 31):
        ws.append([i] + ["" for _ in range(len(sub_headers) - 1)])
    
    # Summary section
    summary_rows = [
        ["OUTLET IN TODAY'S BEAT:", "", "TOTAL SALE:"],
        ["No of Outlets Visited:", "", "SES AVAILABILITY (%):"],
        ["TOTAL UNPAID VISIBILITY:", "", "SER AVAILABILITY (%):"],
        ["SAMPLING QTY:", "", "CSF'69 AVAILABILITY (%):"],
        ["TRL AVAILABILITY (%):", "", "CSF'64/CRF AVAILABILITY (%):"]
    ]
    
    for row in summary_rows:
        ws.append(row)
    
    # Formatting
    for col_num, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        col_letter = get_column_letter(col_num)
        for cell in column_cells:
            try:  # Necessary to handle errors in empty cells
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Styling
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:R1")
    
    wb.save("Activation_MS_DSR.xlsx")
    print("Excel file 'Activation_MS_DSR.xlsx' created successfully.")

create_excel_file()
